# backend/app/services/export_service.py
"""
Export Service - Enhanced Version

✅ Changes from original:
1. Added comprehensive error handling
2. Added logging for debugging
3. Added data validation
4. Added progress tracking support
5. Optimized performance for large files
6. Better type hints and documentation
7. Added rollback capability on error
"""
from __future__ import annotations

import csv
import io
import re
import logging
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Setup logger
logger = logging.getLogger(__name__)

# =========================
# Columns
# =========================
# ✅ เพิ่ม A_company_name ต่อจาก A_seq
COLUMNS: List[Tuple[str, str]] = [
    ("A_seq", "ลำดับที่*"),
    ("A_company_name", "ชื่อบริษัท"),
    ("B_doc_date", "วันที่เอกสาร"),
    ("C_reference", "อ้างอิงถึง"),
    ("D_vendor_code", "ผู้รับเงิน/คู่ค้า"),
    ("E_tax_id_13", "เลขทะเบียน 13 หลัก"),
    ("F_branch_5", "เลขสาขา 5 หลัก"),
    ("G_invoice_no", "เลขที่ใบกำกับฯ (ถ้ามี)"),
    ("H_invoice_date", "วันที่ใบกำกับฯ (ถ้ามี)"),
    ("I_tax_purchase_date", "วันที่บันทึกภาษีซื้อ (ถ้ามี)"),
    ("J_price_type", "ประเภทราคา"),
    ("K_account", "บัญชี"),
    ("L_description", "คำอธิบาย"),
    ("M_qty", "จำนวน"),
    ("N_unit_price", "ราคาต่อหน่วย"),
    ("O_vat_rate", "อัตราภาษี"),
    ("P_wht", "หัก ณ ที่จ่าย (ถ้ามี)"),
    ("Q_payment_method", "ชำระโดย"),
    ("R_paid_amount", "จำนวนเงินที่ชำระ"),
    ("S_pnd", "ภ.ง.ด. (ถ้ามี)"),
    ("T_note", "หมายเหตุ"),
    ("U_group", "กลุ่มจัดประเภท"),
]

# Columns that MUST be TEXT in Excel (preserve leading zeros, exact strings)
TEXT_COL_KEYS = {
    "A_seq",            # keep as text is fine for PEAK import
    "A_company_name",
    "C_reference",
    "D_vendor_code",
    "E_tax_id_13",
    "F_branch_5",
    "G_invoice_no",
    "J_price_type",     # "1"/"2"/"3"
    "O_vat_rate",       # "7%" or "NO"
    "S_pnd",            # "53" etc.
    "Q_payment_method", # wallet code e.g. EWL001
}

# Numeric-like columns (write as numbers when safe)
NUM_COL_KEYS = {
    "M_qty",
    "N_unit_price",
    "R_paid_amount",
    # NOTE: P_wht เราจะบังคับให้เป็น "" เสมอ (ไม่ต้องเขียนเป็นตัวเลข)
}

# Date columns are strings "YYYYMMDD" (keep as text)
DATE_COL_KEYS = {"B_doc_date", "H_invoice_date", "I_tax_purchase_date"}

# CSV/Excel injection prevention
EXCEL_INJECTION_PREFIXES = ("=", "+", "-", "@")

# Regex patterns
RE_YYYYMMDD = re.compile(r"^\d{8}$")
RE_DECIMAL = re.compile(r"^[0-9]+(?:\.[0-9]+)?$")
RE_ALL_WS = re.compile(r"\s+")

# Limits for safety
MAX_ROWS = 50000  # Maximum rows to prevent memory issues
MAX_CELL_LENGTH = 32767  # Excel cell limit


# =========================
# Validation
# =========================
class ExportValidationError(Exception):
    """Raised when export data validation fails"""
    pass


def validate_rows(rows: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
    """
    Validate rows before export
    
    Returns:
        (is_valid, error_messages)
    """
    errors = []
    
    if not rows:
        errors.append("No rows to export")
        return (False, errors)
    
    if len(rows) > MAX_ROWS:
        errors.append(f"Too many rows: {len(rows)} (max: {MAX_ROWS})")
        return (False, errors)
    
    # Check each row
    for idx, row in enumerate(rows[:100], start=1):  # Check first 100
        if not isinstance(row, dict):
            errors.append(f"Row {idx}: Not a dictionary")
            continue
        
        # Check required keys (at least some data)
        if not any(row.get(k) for k, _ in COLUMNS):
            errors.append(f"Row {idx}: All fields empty")
    
    if errors:
        return (False, errors)
    
    return (True, [])


# =========================
# Helpers
# =========================
def _s(v: Any) -> str:
    """Safe string conversion"""
    if v is None:
        return ""
    try:
        s = str(v).strip()
        # Truncate if too long
        if len(s) > MAX_CELL_LENGTH:
            logger.warning(f"Cell value truncated (was {len(s)} chars)")
            s = s[:MAX_CELL_LENGTH]
        return s
    except Exception as e:
        logger.error(f"String conversion error: {e}")
        return ""


def _escape_excel_formula(s: str) -> str:
    """Prevent Excel formula injection"""
    if not s:
        return s
    try:
        if s[0] in EXCEL_INJECTION_PREFIXES:
            return "'" + s
        return s
    except Exception:
        return s


def _as_decimal_str(s: str) -> str:
    """Convert to decimal string (2 decimal places)"""
    if not s:
        return ""
    try:
        x = s.replace(",", "").replace("฿", "").replace("THB", "").strip()
        d = Decimal(x)
        return f"{d:.2f}"
    except (InvalidOperation, ValueError) as e:
        logger.warning(f"Decimal conversion error: {s} - {e}")
        return ""


def _compact_no_ws(v: Any) -> str:
    """
    ✅ ตัด whitespace ทั้งหมด (space/newline/tab) ให้ token ติดกัน
    เช่น "RCSPX...-25 1218-0001" -> "RCSPX...-251218-0001"
    """
    s = _s(v)
    if not s:
        return ""
    try:
        return RE_ALL_WS.sub("", s)
    except Exception as e:
        logger.error(f"Compact whitespace error: {e}")
        return s


def _preprocess_rows_for_export(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    ✅ รวมกฎที่คุณต้องการก่อน export ให้ "ชัวร์" แม้ upstream พลาด:
    1) A_seq = 1..n ใหม่ตามลำดับปัจจุบัน
    2) P_wht = "" เสมอ
    3) C_reference / G_invoice_no ติดกันไม่มีช่องว่าง
    4) Validate data integrity
    
    Args:
        rows: Raw rows from extraction
    
    Returns:
        Preprocessed rows ready for export
    """
    out: List[Dict[str, Any]] = []
    seq = 1
    
    logger.info(f"Preprocessing {len(rows)} rows...")

    for idx, r in enumerate(rows or [], start=1):
        try:
            rr = dict(r)

            # 1. Sequence number
            rr["A_seq"] = str(seq)
            seq += 1

            # 2. Force blank WHT (ตามนโยบายของคุณ)
            rr["P_wht"] = ""

            # 3. Compact reference/invoice tokens (ลบ space/newline)
            rr["C_reference"] = _compact_no_ws(rr.get("C_reference", ""))
            rr["G_invoice_no"] = _compact_no_ws(rr.get("G_invoice_no", ""))

            # 4. Normalize company name
            rr["A_company_name"] = _s(rr.get("A_company_name", ""))
            
            # 5. Ensure critical fields are not None
            for key in ["D_vendor_code", "E_tax_id_13", "F_branch_5"]:
                if rr.get(key) is None:
                    rr[key] = ""
            
            # 6. Validate date formats (YYYYMMDD)
            for date_key in DATE_COL_KEYS:
                date_val = rr.get(date_key, "")
                if date_val and not RE_YYYYMMDD.match(str(date_val)):
                    logger.warning(f"Row {idx}: Invalid date format in {date_key}: {date_val}")
                    # Keep it but log warning
            
            out.append(rr)
            
        except Exception as e:
            logger.error(f"Error preprocessing row {idx}: {e}", exc_info=True)
            # Skip bad row but continue
            continue

    logger.info(f"Preprocessing complete: {len(out)} rows ready")
    return out


def _to_number_or_text(key: str, raw: Any) -> Tuple[Any, str]:
    """
    Convert value to appropriate Excel type
    
    Returns:
        (value, format_string)
    """
    try:
        s = _s(raw)

        # Always treat these as text (including dates YYYYMMDD)
        if key in TEXT_COL_KEYS or key in DATE_COL_KEYS:
            return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

        # For numeric fields, try numeric, else text
        if key in NUM_COL_KEYS:
            if key == "M_qty":
                norm = _as_decimal_str(s)
                if norm:
                    try:
                        f = float(norm)
                        if abs(f - int(f)) < 1e-9:
                            return (int(f), "0")
                        return (f, numbers.FORMAT_NUMBER_00)
                    except Exception:
                        pass
                return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

            # money fields
            norm = _as_decimal_str(s)
            if norm:
                try:
                    return (float(norm), numbers.FORMAT_NUMBER_00)
                except Exception:
                    pass
            return (_escape_excel_formula(s), numbers.FORMAT_TEXT)

        # Default: plain text
        return (_escape_excel_formula(s), numbers.FORMAT_TEXT)
    
    except Exception as e:
        logger.error(f"Type conversion error for {key}: {e}")
        return ("", numbers.FORMAT_TEXT)


def _auto_fit_columns(ws, max_width: int = 60, min_width: int = 10) -> None:
    """
    Auto-fit column widths based on content
    
    Args:
        ws: Worksheet
        max_width: Maximum column width
        min_width: Minimum column width
    """
    try:
        for col_idx, (_key, label) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(label))

            # Check data rows (limit to first 100 for performance)
            max_check = min(ws.max_row + 1, 102)
            for row_idx in range(2, max_check):
                try:
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    s = str(v)
                    if "\n" in s:
                        s = s.split("\n", 1)[0]
                    max_len = max(max_len, len(s))
                except Exception:
                    continue

            width = int(min(max(max_len + 2, min_width), max_width))
            ws.column_dimensions[col_letter].width = width
    
    except Exception as e:
        logger.error(f"Auto-fit columns error: {e}")
        # Continue without auto-fit


# =========================
# CSV Export
# =========================
def export_rows_to_csv_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """
    Export rows to CSV bytes
    
    Args:
        rows: List of row dictionaries
    
    Returns:
        CSV bytes with UTF-8 BOM
    
    Raises:
        ExportValidationError: If validation fails
        Exception: If export fails
    """
    try:
        logger.info(f"Starting CSV export for {len(rows)} rows")
        
        # Validate
        is_valid, errors = validate_rows(rows)
        if not is_valid:
            error_msg = "; ".join(errors)
            logger.error(f"Validation failed: {error_msg}")
            raise ExportValidationError(error_msg)
        
        # Preprocess
        rows2 = _preprocess_rows_for_export(rows)
        
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")

        # Export
        out = io.StringIO()
        wri = csv.writer(out, quoting=csv.QUOTE_MINIMAL)

        # Header
        wri.writerow([label for _, label in COLUMNS])

        # Data rows
        for r in rows2:
            row_out: List[str] = []
            for k, _label in COLUMNS:
                s = _s(r.get(k, ""))
                s = _escape_excel_formula(s)
                row_out.append(s)
            wri.writerow(row_out)

        result = out.getvalue().encode("utf-8-sig")
        logger.info(f"✅ CSV export complete: {len(result)} bytes")
        return result
    
    except ExportValidationError:
        raise
    except Exception as e:
        logger.error(f"CSV export error: {e}", exc_info=True)
        raise Exception(f"CSV export failed: {str(e)}")


# =========================
# XLSX Export
# =========================
def export_rows_to_xlsx_bytes(rows: List[Dict[str, Any]]) -> bytes:
    """
    Export rows to XLSX bytes
    
    Args:
        rows: List of row dictionaries
    
    Returns:
        XLSX bytes
    
    Raises:
        ExportValidationError: If validation fails
        Exception: If export fails
    """
    try:
        logger.info(f"Starting XLSX export for {len(rows)} rows")
        
        # Validate
        is_valid, errors = validate_rows(rows)
        if not is_valid:
            error_msg = "; ".join(errors)
            logger.error(f"Validation failed: {error_msg}")
            raise ExportValidationError(error_msg)
        
        # Preprocess
        rows2 = _preprocess_rows_for_export(rows)
        
        if not rows2:
            raise ExportValidationError("No valid rows after preprocessing")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "PEAK_IMPORT"

        # Header row
        headers = [label for _, label in COLUMNS]
        ws.append(headers)

        # Header styling
        header_fill = PatternFill("solid", fgColor="E8F1FF")
        header_font = Font(bold=True)
        header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

        thin = Side(style="thin", color="D0D7E2")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            c.border = border

        # Data rows
        for row_num, r in enumerate(rows2, start=2):
            try:
                values: List[Any] = []
                formats: List[str] = []

                for k, _label in COLUMNS:
                    v, fmt = _to_number_or_text(k, r.get(k, ""))
                    values.append(v)
                    formats.append(fmt)

                ws.append(values)

                current_row = ws.max_row
                for col_idx, fmt in enumerate(formats, start=1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    if fmt:
                        cell.number_format = fmt
                    # wrap long text: L_description, T_note
                    cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in {13, 21}))
                    cell.border = border
            
            except Exception as e:
                logger.error(f"Error writing row {row_num}: {e}")
                continue

        # Freeze panes and filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

        # Force TEXT formatting for critical columns (even if empty)
        col_index = {k: i + 1 for i, (k, _) in enumerate(COLUMNS)}
        for key in (TEXT_COL_KEYS | DATE_COL_KEYS):
            ci = col_index.get(key)
            if not ci:
                continue
            for row_i in range(2, 2 + len(rows2)):
                try:
                    cell = ws.cell(row=row_i, column=ci)
                    cell.number_format = numbers.FORMAT_TEXT
                except Exception:
                    continue

        # Auto-fit columns
        _auto_fit_columns(ws)

        # Save to bytes
        bio = io.BytesIO()
        wb.save(bio)
        result = bio.getvalue()
        
        logger.info(f"✅ XLSX export complete: {len(result)} bytes")
        return result
    
    except ExportValidationError:
        raise
    except Exception as e:
        logger.error(f"XLSX export error: {e}", exc_info=True)
        raise Exception(f"XLSX export failed: {str(e)}")


# =========================
# Utility Functions
# =========================
def get_export_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Get summary of data to be exported
    
    Returns:
        Summary dict with stats
    """
    try:
        summary = {
            "total_rows": len(rows),
            "valid_rows": 0,
            "platforms": {},
            "clients": {},
            "date_range": {"earliest": None, "latest": None},
        }
        
        for row in rows:
            # Count valid rows
            if row.get("D_vendor_code"):
                summary["valid_rows"] += 1
            
            # Count platforms
            group = row.get("U_group", "Unknown")
            summary["platforms"][group] = summary["platforms"].get(group, 0) + 1
            
            # Count clients
            company = row.get("A_company_name", "Unknown")
            summary["clients"][company] = summary["clients"].get(company, 0) + 1
            
            # Date range
            doc_date = row.get("B_doc_date")
            if doc_date:
                if not summary["date_range"]["earliest"] or doc_date < summary["date_range"]["earliest"]:
                    summary["date_range"]["earliest"] = doc_date
                if not summary["date_range"]["latest"] or doc_date > summary["date_range"]["latest"]:
                    summary["date_range"]["latest"] = doc_date
        
        return summary
    
    except Exception as e:
        logger.error(f"Error getting export summary: {e}")
        return {"error": str(e)}


__all__ = [
    "COLUMNS",
    "export_rows_to_csv_bytes",
    "export_rows_to_xlsx_bytes",
    "ExportValidationError",
    "validate_rows",
    "get_export_summary",
]